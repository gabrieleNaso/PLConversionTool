#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parents[1]
BACKEND_ROOT = PROJECT_ROOT / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.core_converter import export_conversion_bundle_from_ir  # noqa: E402

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - runtime guard
    raise SystemExit(
        "Modulo mancante: openpyxl. Installa le dipendenze backend (pip install -r backend/requirements.txt)."
    ) from exc


def _slugify(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_") or "Sequence"


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _pick(row: dict[str, object], *keys: str) -> object:
    for key in keys:
        if key in row and row[key] is not None and _cell_text(row[key]):
            return row[key]
    for key in keys:
        if key in row:
            return row[key]
    return None


def _split_list(value: object) -> list[str]:
    text = _cell_text(value)
    if not text:
        return []
    return [item.strip() for item in re.split(r"[|,;]", text) if item.strip()]


def _split_int_list(value: object) -> list[int]:
    items: list[int] = []
    for token in _split_list(value):
        match = re.search(r"-?\d+", token)
        if match:
            items.append(int(match.group(0)))
    return items


def _int_or_default(value: object, default: int) -> int:
    text = _cell_text(value)
    if not text:
        return default
    match = re.search(r"-?\d+", text)
    if not match:
        return default
    return int(match.group(0))


def _int_or_none(value: object) -> int | None:
    text = _cell_text(value)
    if not text:
        return None
    match = re.search(r"-?\d+", text)
    if not match:
        return None
    parsed = int(match.group(0))
    return parsed if parsed > 0 else None


def _contains_token(text: str, token: str) -> bool:
    if not text or not token:
        return False
    return re.search(rf"\b{re.escape(token)}\b", text, flags=re.IGNORECASE) is not None


def _normalize_flow_type(value: object) -> str:
    raw = _cell_text(value).strip().lower()
    if raw in {"parallel", "parallelo"}:
        return "parallel"
    return "alternative"


def _normalize_parallel_group(value: object) -> str:
    raw = _cell_text(value).strip()
    return re.sub(r"\s+", "_", raw) if raw else ""


def _normalize_operand_category(value: object) -> str:
    raw = _cell_text(value).strip().lower()
    aliases = {
        "allarme": "alarm",
        "allarmi": "alarm",
        "fault": "alarm",
        "faults": "alarm",
        "alarm": "alarm",
        "aux": "aux",
        "ausiliario": "aux",
        "hmi": "hmi",
        "operatore": "hmi",
        "output": "output",
        "uscita": "output",
        "timer": "timer",
        "tempo": "timer",
        "memory": "memory",
        "memoria": "memory",
        "external": "external",
        "esterno": "external",
        "manual_mode": "manual_mode",
        "manual": "manual_mode",
        "auto_mode": "auto_mode",
        "auto": "auto_mode",
    }
    return aliases.get(raw, raw or "aux")


def _normalize_timer_kind(value: object) -> str:
    raw = _cell_text(value).strip().lower()
    if not raw:
        return "t_on"
    aliases = {
        "t_on": "t_on",
        "ton": "t_on",
        "sd": "t_on",
        "t_off": "t_off",
        "tof": "t_off",
        "sf": "t_off",
        "t_p": "t_p",
        "tp": "t_p",
        "se": "t_p",
        "sp": "t_p",
        "ss": "t_p",
    }
    return aliases.get(raw, "t_on")


def _normalize_plc_datatype(value: object) -> str:
    raw = _cell_text(value).strip().lower()
    if not raw:
        return "Bool"
    aliases = {
        "bool": "Bool",
        "boolean": "Bool",
        "int": "Int",
        "integer": "Int",
        "dint": "DInt",
        "udint": "UDInt",
        "real": "Real",
        "byte": "Byte",
        "word": "Word",
        "dword": "DWord",
        "time": "Time",
        "timer": "IEC_TIMER",
        "iec_timer": "IEC_TIMER",
        "string": "String",
    }
    return aliases.get(raw, _cell_text(value).strip() or "Bool")


def _normalize_support_category(value: object) -> str:
    raw = _cell_text(value).strip().lower()
    aliases = {
        "io": "io",
        "ingressi": "io",
        "inputs": "io",
        "diag": "diag",
        "diagnostica": "diag",
        "mode": "mode",
        "modalita": "mode",
        "modalità": "mode",
        "transitions": "transitions",
        "transition": "transitions",
        "transizioni": "transitions",
        "output": "output",
        "uscite": "output",
        "external": "external",
        "ext": "external",
        "esterno": "external",
        "hmi": "hmi",
        "aux": "aux",
    }
    return aliases.get(raw, "")


def _read_support_members(path: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for sheet_name in ("support_fc", "fc_support"):
        rows.extend(_read_sheet_rows(path, sheet_name))

    support_members: list[dict[str, object]] = []
    for row in rows:
        category = _normalize_support_category(
            _pick(row, "category", "categoria", "support_category", "fc_category")
        )
        member_name = _cell_text(_pick(row, "member_name", "name", "operand", "tag", "variabile"))
        if not category or not member_name:
            continue
        support_members.append(
            {
                "category": category,
                "member_name": member_name,
                "comment": _cell_text(_pick(row, "comment", "note", "notes", "description", "descrizione")),
                "network_index": _int_or_none(_pick(row, "network", "network_index", "network_no", "rete")),
                "network_title": _cell_text(_pick(row, "network_title", "title", "network_name")),
            }
        )
    return support_members


def _read_support_logic_rows(path: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    # Primary format: logic rows on the same support_fc sheet.
    # Legacy aliases are still accepted for backward compatibility.
    for sheet_name in ("support_fc", "fc_support", "support_fc_logic", "fc_logic_support", "fc_logic"):
        rows.extend(_read_sheet_rows(path, sheet_name))

    logic_rows: list[dict[str, object]] = []
    for row in rows:
        category = _normalize_support_category(
            _pick(row, "category", "categoria", "support_category", "fc_category")
        )
        result_member = _cell_text(
            _pick(row, "result_member", "coil_member", "target_member", "member_name", "output_member")
        )
        if not category or not result_member:
            continue
        condition_expression = _cell_text(
            _pick(row, "condition_expression", "guard_expression", "expression", "condition")
        )
        condition_operands = _split_list(
            _pick(row, "condition_operands", "guard_operands", "operands", "input_operands")
        )
        if not condition_expression and condition_operands:
            condition_expression = " AND ".join(condition_operands)
        logic_rows.append(
            {
                "category": category,
                "network_index": _int_or_none(_pick(row, "network", "network_index", "network_no", "rete")),
                "network_title": _cell_text(_pick(row, "network_title", "title", "network_name")),
                "result_member": result_member,
                "condition_expression": condition_expression or "TRUE",
                "condition_operands": condition_operands,
                "comment": _cell_text(_pick(row, "comment", "note", "notes", "description", "descrizione")),
            }
        )
    return logic_rows


def _ensure_required_excel_sections(
    *,
    operand_rows: list[dict[str, object]],
    support_members: list[dict[str, object]],
    support_logic: list[dict[str, object]],
) -> None:
    if not operand_rows:
        raise SystemExit(
            "Excel non valido: il foglio 'operands' e' obbligatorio e deve contenere almeno una riga compilata."
        )
    if not support_members and not support_logic:
        raise SystemExit(
            "Excel non valido: il foglio 'support_fc' (o alias 'fc_support') e' obbligatorio e deve contenere almeno una riga compilata (member_name e/o result_member)."
        )


def _dedupe_dict_rows(items: list[dict[str, object]], keys: tuple[str, ...]) -> list[dict[str, object]]:
    seen: set[tuple[str, ...]] = set()
    deduped: list[dict[str, object]] = []
    for item in items:
        signature = tuple(_cell_text(item.get(key)) for key in keys)
        if signature in seen:
            continue
        seen.add(signature)
        deduped.append(item)
    return deduped


def _infer_network_index_for_operand(operand: str, transitions: list[dict[str, object]], default: int) -> int:
    if not operand:
        return default
    for transition in transitions:
        if operand in transition.get("guard_operands", []):
            return int(transition.get("network_index", 0) or default)
        if _contains_token(str(transition.get("guard_expression", "")), operand):
            return int(transition.get("network_index", 0) or default)
    return default



def _read_sheet_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [_cell_text(item) for item in values[0]]
    rows: list[dict[str, object]] = []
    for row in values[1:]:
        if not any(_cell_text(item) for item in row):
            continue
        item = {header: cell for header, cell in zip(headers, row) if header}
        if item:
            rows.append(item)
    return rows


def _read_meta(path: Path) -> dict[str, str]:
    rows = _read_sheet_rows(path, "meta")
    meta: dict[str, str] = {}
    for row in rows:
        key = _cell_text(row.get("key")).lower()
        value = _cell_text(row.get("value"))
        if key:
            meta[key] = value
    return meta


def _build_steps_from_sequence_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    step_map: dict[str, dict[str, object]] = {}
    for row in rows:
        explicit_name = _cell_text(_pick(row, "step_name", "name", "step"))
        explicit_no = _int_or_none(_pick(row, "numero_step", "step_number", "step_no", "numero", "number"))
        if explicit_name:
            step_map.setdefault(
                explicit_name,
                {
                    "name": explicit_name,
                    "step_number": explicit_no,
                    "source_networks": [],
                    "activation_networks": [],
                    "action_networks": [],
                },
            )
            if step_map[explicit_name].get("step_number") is None and explicit_no is not None:
                step_map[explicit_name]["step_number"] = explicit_no

        from_name = _cell_text(_pick(row, "from_step", "source_step"))
        from_no = _int_or_none(_pick(row, "from_step_number"))
        if from_name:
            step_map.setdefault(
                from_name,
                {
                    "name": from_name,
                    "step_number": from_no,
                    "source_networks": [],
                    "activation_networks": [],
                    "action_networks": [],
                },
            )
            if step_map[from_name].get("step_number") is None and from_no is not None:
                step_map[from_name]["step_number"] = from_no

        to_name = _cell_text(_pick(row, "to_step", "target_step"))
        to_no = _int_or_none(_pick(row, "to_step_number"))
        if to_name:
            step_map.setdefault(
                to_name,
                {
                    "name": to_name,
                    "step_number": to_no,
                    "source_networks": [],
                    "activation_networks": [],
                    "action_networks": [],
                },
            )
            if step_map[to_name].get("step_number") is None and to_no is not None:
                step_map[to_name]["step_number"] = to_no
    return list(step_map.values())


def _build_transitions_from_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    transitions: list[dict[str, object]] = []
    for idx, row in enumerate(rows, start=1):
        source_step = _cell_text(_pick(row, "from_step", "source_step"))
        target_step = _cell_text(_pick(row, "to_step", "target_step"))
        if not source_step or not target_step:
            continue
        # Network index is always auto-managed by generator.
        network_index = idx
        guard_expression = _cell_text(_pick(row, "condition_expression", "guard_expression")) or "TRUE"
        guard_operands = _split_list(_pick(row, "operands_used_in_condition", "guard_operands"))
        if not guard_operands and guard_expression and guard_expression.upper() != "TRUE":
            guard_operands = [
                token
                for token in re.findall(r"[A-Za-z_]\w*(?:\.\w+)*", guard_expression)
                if token.upper() not in {"AND", "OR", "NOT", "TRUE", "FALSE"}
            ]
        transitions.append(
            {
                "transition_id": _cell_text(_pick(row, "transition_id")) or f"T{idx}",
                "source_step": source_step,
                "target_step": target_step,
                "network_index": network_index,
                "guard_expression": guard_expression,
                "guard_operands": guard_operands,
                "flow_type": _normalize_flow_type(
                    _pick(row, "flow_type", "branch_mode", "parallel_mode")
                ),
                "parallel_group": _normalize_parallel_group(
                    _pick(row, "parallel_group", "parallel_id", "group_id")
                ),
            }
        )
    return transitions


def build_ir_from_excel(path: Path, sequence_name: str | None = None) -> tuple[str, str, dict]:
    meta = _read_meta(path)
    base_name = sequence_name or meta.get("sequence_name") or path.stem
    normalized_sequence = _slugify(base_name)
    source_name = meta.get("source_name") or path.name

    explicit_networks: list[dict[str, object]] = []
    # Excel input does not expose manual network modeling.
    # Keep list empty and synthesize it from transitions/operands.

    sequence_rows = _read_sheet_rows(path, "sequence")
    if sequence_rows:
        steps = _build_steps_from_sequence_rows(sequence_rows)
        transitions = _build_transitions_from_rows(sequence_rows)
    else:
        steps = []
        for row in _read_sheet_rows(path, "steps"):
            name = _cell_text(_pick(row, "step_name", "name"))
            if not name:
                continue
            steps.append(
                {
                    "name": name,
                    "step_number": _int_or_none(
                        _pick(row, "numero_step", "step_number", "step_no", "numero", "number")
                    ),
                    "source_networks": _split_int_list(
                        _pick(
                            row,
                            "networks_where_step_is_read",
                            "source_networks",
                        )
                    ),
                    "activation_networks": _split_int_list(
                        _pick(
                            row,
                            "networks_where_step_is_activated",
                            "activation_networks",
                        )
                    ),
                    "action_networks": _split_int_list(
                        _pick(
                            row,
                            "networks_with_step_actions",
                            "action_networks",
                        )
                    ),
                }
            )
        transitions = _build_transitions_from_rows(_read_sheet_rows(path, "transitions"))

    timers: list[dict[str, object]] = []
    timer_rows = _read_sheet_rows(path, "timers")
    for idx, row in enumerate(timer_rows, start=1):
        source_timer = _cell_text(_pick(row, "timer_name", "source_timer"))
        if not source_timer:
            continue
        inferred_network = None
        for transition in transitions:
            if source_timer in transition.get("guard_operands", []):
                inferred_network = int(transition.get("network_index", 0) or 0)
                break
            if _contains_token(str(transition.get("guard_expression", "")), source_timer):
                inferred_network = int(transition.get("network_index", 0) or 0)
                break
        timers.append(
            {
                "source_timer": source_timer,
                "network_index": inferred_network or idx,
                "kind": _normalize_timer_kind(_pick(row, "timer_instruction_kind", "kind")),
                "preset": _cell_text(
                    _pick(row, "timer_preset_value", "preset")
                ) or None,
                "trigger_operands": _split_list(
                    _pick(row, "timer_trigger_operands", "trigger_operands")
                ),
            }
        )

    memories: list[dict[str, object]] = []
    memory_rows = _read_sheet_rows(path, "memories")
    for idx, row in enumerate(memory_rows, start=1):
        name = _cell_text(_pick(row, "memory_operand", "name"))
        if not name:
            continue
        inferred_network = None
        for transition in transitions:
            if name in transition.get("guard_operands", []):
                inferred_network = int(transition.get("network_index", 0) or 0)
                break
            if _contains_token(str(transition.get("guard_expression", "")), name):
                inferred_network = int(transition.get("network_index", 0) or 0)
                break
        memories.append(
            {
                "name": name,
                "role": _cell_text(_pick(row, "memory_role", "role")) or "aux",
                "network_index": inferred_network or idx,
            }
        )

    faults: list[dict[str, object]] = []
    fault_rows = _read_sheet_rows(path, "faults")
    for idx, row in enumerate(fault_rows, start=1):
        name = _cell_text(_pick(row, "fault_tag", "name"))
        if not name:
            continue
        inferred_network = None
        for transition in transitions:
            if _contains_token(str(transition.get("guard_expression", "")), name):
                inferred_network = int(transition.get("network_index", 0) or 0)
                break
        faults.append(
            {
                "name": name,
                "network_index": inferred_network or idx,
                "evidence": _cell_text(_pick(row, "fault_evidence", "evidence")),
            }
        )

    outputs: list[dict[str, object]] = []
    output_rows = _read_sheet_rows(path, "outputs")
    for idx, row in enumerate(output_rows, start=1):
        name = _cell_text(_pick(row, "output_operand", "name"))
        if not name:
            continue
        inferred_network = None
        for transition in transitions:
            if _contains_token(str(transition.get("guard_expression", "")), name):
                inferred_network = int(transition.get("network_index", 0) or 0)
                break
        outputs.append(
            {
                "name": name,
                "network_index": inferred_network or idx,
                "action": "=",
            }
        )

    manual_logic_networks = _split_int_list(meta.get("manual_logic_networks"))
    auto_logic_networks = _split_int_list(meta.get("auto_logic_networks"))
    external_refs = _split_list(meta.get("external_refs"))
    support_members = _read_support_members(path)
    support_logic = _read_support_logic_rows(path)

    # New readable format: optional operand catalog that classifies signals
    # used in LAD transition logic by function (alarm/aux/hmi/output/timer/...).
    operand_rows = _read_sheet_rows(path, "operands")
    _ensure_required_excel_sections(
        operand_rows=operand_rows,
        support_members=support_members,
        support_logic=support_logic,
    )
    operand_catalog: list[str] = []
    operand_datatypes: dict[str, str] = {}
    for idx, row in enumerate(operand_rows, start=1):
        operand = _cell_text(_pick(row, "operand", "name", "tag"))
        if not operand:
            continue
        if operand not in operand_catalog:
            operand_catalog.append(operand)
        datatype = _normalize_plc_datatype(
            _pick(row, "datatype", "data_type", "plc_datatype", "plc_type", "tipo_dato")
        )
        if operand not in operand_datatypes:
            operand_datatypes[operand] = datatype
        category = _normalize_operand_category(_pick(row, "category", "type", "group"))
        network_index = _infer_network_index_for_operand(operand, transitions, idx)
        note = _cell_text(_pick(row, "note", "notes", "evidence"))

        if category == "alarm":
            faults.append(
                {
                    "name": operand,
                    "network_index": network_index,
                    "evidence": note or "Classified from operands sheet",
                }
            )
            continue
        if category == "output":
            outputs.append(
                {
                    "name": operand,
                    "network_index": network_index,
                    "action": "=",
                }
            )
            continue
        if category == "timer":
            timers.append(
                {
                    "source_timer": operand,
                    "network_index": network_index,
                    "kind": _normalize_timer_kind(_pick(row, "timer_instruction_kind", "kind")),
                    "preset": _cell_text(_pick(row, "timer_preset_value", "preset")) or None,
                    "trigger_operands": _split_list(_pick(row, "trigger_operands", "timer_trigger_operands")),
                }
            )
            continue
        if category == "hmi":
            external_refs.append(operand)
            memories.append(
                {
                    "name": operand,
                    "role": "hmi",
                    "network_index": network_index,
                }
            )
            continue
        if category == "external":
            external_refs.append(operand)
            continue
        if category == "manual_mode":
            if network_index not in manual_logic_networks:
                manual_logic_networks.append(network_index)
            continue
        if category == "auto_mode":
            if network_index not in auto_logic_networks:
                auto_logic_networks.append(network_index)
            continue

        # Default mapping for aux/memory/other custom categories.
        memories.append(
            {
                "name": operand,
                "role": category if category != "memory" else "aux",
                "network_index": network_index,
            }
        )

    timers = _dedupe_dict_rows(timers, ("source_timer", "network_index"))
    memories = _dedupe_dict_rows(memories, ("name", "network_index"))
    faults = _dedupe_dict_rows(faults, ("name", "network_index"))
    outputs = _dedupe_dict_rows(outputs, ("name", "network_index", "action"))
    external_refs = sorted(set(external_refs))
    manual_logic_networks = sorted(set(manual_logic_networks))
    auto_logic_networks = sorted(set(auto_logic_networks))

    # Step references in transitions are normalized using explicit step_number
    # mapping, so Excel naming stays free and consistent.
    explicit_by_number: dict[int, str] = {}
    explicit_names = {
        str(item.get("name") or "").strip()
        for item in steps
        if str(item.get("name") or "").strip() and item.get("step_number") is not None
    }
    for item in steps:
        name = str(item.get("name") or "").strip()
        number = item.get("step_number")
        if not name or number is None:
            continue
        try:
            step_no = int(number)
        except (TypeError, ValueError):
            continue
        if step_no > 0 and step_no not in explicit_by_number:
            explicit_by_number[step_no] = name

    alias_pattern = re.compile(r"^S(\d+)$", flags=re.IGNORECASE)

    def _normalize_step_ref(token: str) -> str:
        token = token.strip()
        match = alias_pattern.match(token)
        if not match:
            return token
        if token in explicit_names:
            return token
        step_no = int(match.group(1))
        return explicit_by_number.get(step_no, token)

    for transition in transitions:
        source = str(transition.get("source_step") or "").strip()
        target = str(transition.get("target_step") or "").strip()
        if source:
            transition["source_step"] = _normalize_step_ref(source)
        if target:
            transition["target_step"] = _normalize_step_ref(target)

    # Remove alias-only steps (e.g. "S1") when a canonical explicit step with
    # the same number exists (e.g. "Init" with step_number=1).
    canonical_by_number = {number: name for number, name in explicit_by_number.items() if name}
    cleaned_steps: list[dict[str, object]] = []
    for item in steps:
        name = str(item.get("name") or "").strip()
        step_number = item.get("step_number")
        match = alias_pattern.match(name)
        if match and step_number is None:
            alias_no = int(match.group(1))
            canonical_name = canonical_by_number.get(alias_no)
            if canonical_name and canonical_name != name:
                continue
        cleaned_steps.append(item)
    steps = cleaned_steps

    # Enrich step metadata from transitions when not explicitly provided in Excel.
    step_map: dict[str, dict[str, object]] = {item["name"]: item for item in steps}
    for transition in transitions:
        source = str(transition.get("source_step") or "")
        target = str(transition.get("target_step") or "")
        network_index = int(transition.get("network_index", 0) or 0)
        if not network_index:
            continue
        if source in step_map:
            current = set(step_map[source].get("source_networks") or [])
            current.add(network_index)
            step_map[source]["source_networks"] = sorted(current)
        if target in step_map:
            current = set(step_map[target].get("activation_networks") or [])
            current.add(network_index)
            step_map[target]["activation_networks"] = sorted(current)
            actions = set(step_map[target].get("action_networks") or [])
            actions.add(network_index)
            step_map[target]["action_networks"] = sorted(actions)
    steps = list(step_map.values())

    # Build synthetic network registry when the sheet is not used.
    inferred_network_ids = {
        int(item.get("network_index", 0) or 0)
        for item in (transitions + timers + memories + faults + outputs)
        if int(item.get("network_index", 0) or 0) > 0
    }
    inferred_networks = [
        {"index": network_id, "title": f"Excel_Network_{network_id}", "raw_lines": []}
        for network_id in sorted(inferred_network_ids)
    ]
    networks = explicit_networks or inferred_networks

    ir_payload = {
        "sequence_name": normalized_sequence,
        "source_name": source_name,
        "networks": networks,
        "steps": steps,
        "transitions": transitions,
        "timers": timers,
        "memories": memories,
        "faults": faults,
        "outputs": outputs,
        "manual_logic_networks": manual_logic_networks,
        "auto_logic_networks": auto_logic_networks,
        "external_refs": external_refs,
        "strict_operand_catalog": True,
        "operand_catalog": operand_catalog,
        "operand_datatypes": operand_datatypes,
        "support_members": support_members,
        "support_logic": support_logic,
        "assumptions": _split_list(meta.get("assumptions")),
    }
    return normalized_sequence, source_name, ir_payload


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate XML bundle from a manual Excel IR workbook (.xlsx)."
    )
    parser.add_argument("--excel", required=True, help="Path to Excel workbook (.xlsx).")
    parser.add_argument(
        "--output-root",
        default="data/output/generated",
        help="Output root folder (default: data/output/generated).",
    )
    parser.add_argument(
        "--sequence-name",
        default=None,
        help="Optional override for sequence name.",
    )
    parser.add_argument(
        "--keep-existing",
        action="store_true",
        help="Do not clean target bundle folder before generation.",
    )
    args = parser.parse_args()

    excel_path = (PROJECT_ROOT / args.excel).resolve()
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")

    # Keep internal sequence/block naming aligned with package folder name.
    # Default source for both is the Excel filename (sanitized), unless explicitly overridden.
    effective_sequence_name = args.sequence_name or _slugify(excel_path.stem)
    sequence_name, source_name, ir_payload = build_ir_from_excel(
        path=excel_path,
        sequence_name=effective_sequence_name,
    )

    output_root = (PROJECT_ROOT / args.output_root).resolve()
    # Keep package folder aligned with internal sequence/block naming.
    bundle_dir = output_root / sequence_name
    if bundle_dir.exists() and not args.keep_existing:
        shutil.rmtree(bundle_dir)
    bundle_dir.mkdir(parents=True, exist_ok=True)
    bundle_relative = bundle_dir.relative_to(PROJECT_ROOT)

    ir_json_path = bundle_dir / f"{sequence_name}_ir.json"
    ir_json_path.write_text(json.dumps(ir_payload, indent=2), encoding="utf-8")

    result = export_conversion_bundle_from_ir(
        sequence_name=sequence_name,
        ir_payload=ir_payload,
        source_name=source_name,
        output_dir=str(bundle_relative),
    )
    print(f"[OK] {excel_path.name} -> {result['outputDirectory']}")
    print(f"[IR] {ir_json_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
